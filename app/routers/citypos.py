import datetime
from io import BytesIO
from typing import Dict, Any
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, status
from fastapi.responses import StreamingResponse
import pandas as pd
from app.database import conn, cursor
import traceback

from app.utils.auth import get_current_user
from app.utils.logger import log_user_action
from app.utils.role_check import check_role

router = APIRouter()

# ================================
# Utility Function
# ================================
def _truncate(value: Any, max_len: int = 255, column_type: str = "text"):
    """
    Clean & safely truncate data values.
    Returns None for empty / invalid dates so that PostgreSQL gets NULL.
    """
    # 1. treat pandas NaT / NaN / None / empty string as “missing”
    if pd.isna(value) or str(value).strip() in ("", "NaT"):
        return None if column_type == "date" else " "

    s = str(value).strip()

    if column_type == "date":
        try:
            # convert -> real date object -> NULL on failure
            return pd.to_datetime(s).date()
        except Exception:
            return None
    # text
    return s[:max_len] if s else " "


# ================================
# 1️⃣ Get all city data
# ================================
@router.get("/city/all", status_code=status.HTTP_200_OK)
def get_all_city(current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["support", "admin", "superadmin"])
    try:
        cursor.execute('SELECT * FROM city ORDER BY id ASC;')
        rows = cursor.fetchall()
        return {"data": rows}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



# ================================
# 4️⃣ Upload Excel
# ================================
@router.post("/city/upload", status_code=status.HTTP_201_CREATED)
def upload_city_excel(file: UploadFile = File(...),current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["admin", "superadmin"])
    try:
        df = pd.read_excel(BytesIO(file.file.read()))
        df.columns = [c.strip() for c in df.columns]
        df = df.fillna("").dropna(how="all")

        db_cols = [
            "SL", "PURPOSE", "CONFIGDATE", "OLD MID", "OLD TID", "MID", "TID",
            "MERCHANT NAME", "DBA NAME", "ADDRESS", "CITY", "LOCATION", "VENDOR",
            "POS TYPE", "POS MODEL", "POS SERIAL", "OLD POS SERIAL",
            "OLD SIM SERIAL", "SIM SERIAL NUMBER", "SIM NUMBER", "SIM CARRIER",
            "IP ADDRESS", "PORT NUMBER", "REASON", "SPECIAL FUNCTIONALITY",
            "INSTALLATION DATE", "MERCHAN CONTACT PERSON",
            "MERCHAN CONTACT NUMBER", "HANDOVER TO", "HANDOVER DATE",
            "ROLL OUT BY", "ROLL OUT DATE", "APP RELEASE DATE", "QR CODE", "QR ID"
        ]

        date_columns = [
            "INSTALLATION DATE", "HANDOVER DATE", "ROLL OUT DATE", "APP RELEASE DATE"
        ]

        inserted = updated = skipped = failed = 0

        for _, row in df.iterrows():
            try:
                record = {}
                for col in df.columns:
                    if col in db_cols:
                        col_type = "date" if col in date_columns else "text"
                        record[col] = _truncate(row[col], column_type=col_type)

                pos_serial = record.get("POS SERIAL")
                if not pos_serial:
                    skipped += 1
                    continue

                cursor.execute('SELECT id FROM city WHERE "POS SERIAL" = %s;', (pos_serial,))
                exists = cursor.fetchone()

                if exists:
                    set_clause = ", ".join([f'"{k}" = %s' for k in record])
                    cursor.execute(
                        f'UPDATE city SET {set_clause}, update_time = NOW() WHERE "POS SERIAL" = %s;',

                        (*record.values(), pos_serial)
                    )
                    updated += 1
                else:
                    placeholders = ", ".join(["%s"] * len(record))
                    columns = ", ".join([f'"{k}"' for k in record.keys()])
                    cursor.execute(
                        f'INSERT INTO city ({columns}, create_time, update_time) VALUES ({placeholders}, NOW(), NOW());',

                        tuple(record.values())
                    )
                    inserted += 1
            except Exception:
                traceback.print_exc()
                conn.rollback()
                failed += 1

        conn.commit()

        # Log user action
        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="upload_mtb_bulk",
            target_table="Citybank",
            description=f"Bulk uploaded Citybank Excel: {inserted} inserted, {updated} updated, {skipped} skipped, {failed} failed"
        )
        return {
            "message": "✅ Upload completed",
            "summary": {"inserted": inserted, "updated": updated, "failed": failed, "skipped": skipped}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        file.file.close()


# ================================
# 5️⃣ Download Excel
# ================================
@router.get("/city/download", status_code=status.HTTP_200_OK)
def download_all_city_records():
    try:
        columns = [
            "SL", "PURPOSE", "CONFIGDATE", "OLD MID", "OLD TID", "MID", "TID",
            "MERCHANT NAME", "DBA NAME", "ADDRESS", "CITY", "LOCATION", "VENDOR",
            "POS TYPE", "POS MODEL", "POS SERIAL", "OLD POS SERIAL",
            "OLD SIM SERIAL", "SIM SERIAL NUMBER", "SIM NUMBER", "SIM CARRIER",
            "IP ADDRESS", "PORT NUMBER", "REASON", "SPECIAL FUNCTIONALITY",
            "INSTALLATION DATE", "MERCHAN CONTACT PERSON",
            "MERCHAN CONTACT NUMBER", "HANDOVER TO", "HANDOVER DATE",
            "ROLL OUT BY", "ROLL OUT DATE", "APP RELEASE DATE", "QR CODE", "QR ID"
        ]

        quoted_columns = [f'"{c}"' for c in columns]
        cursor.execute(f'SELECT {", ".join(quoted_columns)} FROM city ORDER BY id;')
        rows = cursor.fetchall()
        df = pd.DataFrame(rows, columns=columns)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_name = "City_POS_Data"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            worksheet = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns, 1):
                max_length = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)
                for cell in worksheet[get_column_letter(i)]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        output.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="city_pos_data.xlsx"'}
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {e}")


# ================================
# 6️⃣ Download Template
# ================================
@router.get("/city/template")
def download_city_template():
    cols = [
        "SL", "PURPOSE", "CONFIGDATE", "OLD MID", "OLD TID", "MID", "TID",
        "MERCHANT NAME", "DBA NAME", "ADDRESS", "CITY", "LOCATION", "VENDOR",
        "POS TYPE", "POS MODEL", "POS SERIAL", "OLD POS SERIAL",
        "OLD SIM SERIAL", "SIM SERIAL NUMBER", "SIM NUMBER", "SIM CARRIER",
        "IP ADDRESS", "PORT NUMBER", "REASON", "SPECIAL FUNCTIONALITY",
        "INSTALLATION DATE", "MERCHAN CONTACT PERSON", "MERCHAN CONTACT NUMBER",
        "HANDOVER TO", "HANDOVER DATE", "ROLL OUT BY", "ROLL OUT DATE",
        "APP RELEASE DATE", "QR CODE", "QR ID"
    ]
    df = pd.DataFrame(columns=cols)
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="city_template.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
