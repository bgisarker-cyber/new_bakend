
import datetime
from io import BytesIO
from typing import Dict, Any
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, status
from fastapi.responses import StreamingResponse
import pandas as pd
from app.database import conn, cursor
import traceback
from app.utils.auth import get_current_user
from app.utils.logger import log_user_action
from app.utils.role_check import check_role

router = APIRouter(prefix="/pubali", tags=['pubali'])

# ================================
# Utility
# ================================
def _truncate(value: Any, max_len: int = 255):
    if value is None or str(value).strip() == "":
        return "N/A"
    s = str(value).strip()
    return s[:max_len]


# ================================
# 1️⃣ Get all
# ================================
@router.get("/all", status_code=status.HTTP_200_OK)
def get_all_pubali(current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["support", "admin", "superadmin"])    

    try:
        cursor.execute("SELECT * FROM pubali ORDER BY id ASC;")
        rows = cursor.fetchall()
        return {"data": rows}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ================================
# 2️⃣ Edit
# ================================



# ================================
# 4️⃣ Upload Excel
# ================================
@router.post("/upload", status_code=status.HTTP_201_CREATED)
def upload_pubali_excel(file: UploadFile = File(...),current_user: dict = Depends(get_current_user)):
    check_role(current_user, [ "admin", "superadmin"]) 
    try:
        df = pd.read_excel(BytesIO(file.file.read()))
        df.columns = [c.strip().upper() for c in df.columns]

        if "POS S/N" not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'POS S/N' column in Excel")

        df = df.fillna("").dropna(how="all")

        # Database columns (must match exactly)
        db_cols = {
            "S/N", "CONFIGURATION DATE", "TID", "MID", "MERCHANT NAME",
            "DBA NAME", "ADDRESS", "BANK CONT. PERSON NAME", "BANK CONT. PERSON NUMBER",
            "DISTRICT", "DIVISION", "ZONE", "POS S/N", "POS BRAND & MODEL",
            "OPERATOR NAME", "SIM NUMBER", "SIM IP", "HOST IP", "HOST PORT",
            "EMI", "VENDOR NAME", "EMI TENURE", "REMARKS", "APPS VERSION",
            "CONFIGURE BY", "INSTALLATION DATE", "INSTALLATION BY",
            "MERCHANT NUMBER", "FIRMWARE VERSION", "QR"
        }

        # Helper to truncate string values
        def _truncate(value: str, max_len: int = 255):
            if value is None:
                return None
            return str(value)[:max_len]

        inserted = updated = skipped = failed = 0

        for _, row in df.iterrows():
            try:
                record = {}
                for col in df.columns:
                    if col in db_cols:
                        record[col] = _truncate(row[col])

                pos_serial = record.get("POS S/N")
                if not pos_serial:
                    skipped += 1
                    continue

                # Check if record exists
                cursor.execute('SELECT id FROM pubali WHERE "POS S/N" = %s;', (pos_serial,))
                exists = cursor.fetchone()

                if exists:
                    set_clause = ", ".join([f'"{k}" = %s' for k in record])
                    cursor.execute(
                        f'UPDATE pubali SET {set_clause}, update_time = NOW() WHERE "POS S/N" = %s;',
                        (*record.values(), pos_serial)
                    )
                    updated += 1
                else:
                    placeholders = ", ".join(["%s"] * len(record))
                    columns = ", ".join([f'"{k}"' for k in record.keys()])
                    cursor.execute(
                        f'INSERT INTO pubali ({columns}, create_time, update_time) '
                        f'VALUES ({placeholders}, NOW(), NOW());',
                        tuple(record.values())
                    )
                    inserted += 1
            except Exception as e:
                traceback.print_exc()
                conn.rollback()
                failed += 1

        conn.commit()

        # Log user action
        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="Pubali_bulk",
            target_table="PBL",
            description=f"Bulk uploaded PBL Excel: {inserted} inserted, {updated} updated, {skipped} skipped, {failed} failed"
        )
        return {
            "message": "✅ Upload completed",
            "summary": {"inserted": inserted, "updated": updated, "failed": failed, "skipped": skipped}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        file.file.close()



# ================================
# 6️⃣ Export Excel
# ================================
@router.get("/download", status_code=status.HTTP_200_OK)
def download_all_pubali_records(current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["support", "admin", "superadmin"]) 
    try:
        columns = [
            "S/N", "CONFIGURATION DATE", "TID", "MID", "MERCHANT NAME",
            "DBA NAME", "ADDRESS", "BANK CONT. PERSON NAME", "BANK CONT. PERSON NUMBER",
            "DISTRICT", "DIVISION", "ZONE", "POS S/N", "POS BRAND & MODEL",
            "OPERATOR NAME", "SIM NUMBER", "SIM IP", "HOST IP", "HOST PORT",
            "EMI", "VENDOR NAME", "EMI TENURE", "REMARKS", "APPS VERSION",
            "CONFIGURE BY", "INSTALLATION DATE", "INSTALLATION BY",
            "MERCHANT NUMBER", "FIRMWARE VERSION", "QR"
        ]

        quoted_columns = [f'"{c}"' for c in columns]
        cursor.execute(f'SELECT {", ".join(quoted_columns)} FROM pubali ORDER BY id;')
        rows = cursor.fetchall()
        df = pd.DataFrame(rows, columns=columns)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_name = "Pubali_POS_Data"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # Access workbook and sheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Auto-fit column widths and enable text wrapping
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            for i, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                # Limit to avoid extremely wide columns
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width

                # Apply text wrapping to potentially long fields
                if col.lower() in ["address", "remarks", "merchant_signboard"]:
                    for cell in worksheet[get_column_letter(i)]:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        output.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="pubali_pos_data.xlsx"'}
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {e}")
    



# Router to create template

@router.get("/template")
def download_pubali_template():
    cols = [
        "S/N", "CONFIGURATION DATE", "TID", "MID", "MERCHANT NAME",
        "DBA NAME", "ADDRESS", "BANK CONT. PERSON NAME", "BANK CONT. PERSON NUMBER",
        "DISTRICT", "DIVISION", "ZONE", "POS S/N", "POS BRAND & MODEL",
        "OPERATOR NAME", "SIM NUMBER", "SIM IP", "HOST IP", "HOST PORT",
        "EMI", "VENDOR NAME", "EMI TENURE", "REMARKS", "APPS VERSION",
        "CONFIGURE BY", "INSTALLATION DATE", "INSTALLATION BY",
        "MERCHANT NUMBER", "FIRMWARE VERSION", "QR"
    ]
    df = pd.DataFrame(columns=cols)
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="pubali_template.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)