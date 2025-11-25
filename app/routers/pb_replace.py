import datetime
from io import BytesIO
from typing import Dict, Any
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, status
from fastapi.responses import StreamingResponse
import pandas as pd
from app.database import conn, cursor
import traceback
from app.utils.auth import get_current_user
from app.utils.logger import log_user_action
from app.utils.role_check import check_role

router = APIRouter(prefix="/pubali_replace", tags=["pubali_replace"])

# ================================
# Helper Function
# ================================
def _truncate(value: Any, max_len: int = 255):
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip()[:max_len]

# ================================
# 1️⃣ Get All Records
# ================================
@router.get("/all", status_code=status.HTTP_200_OK)
def get_all_pubali_replace(current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["support", "admin", "superadmin"])
    try:
        cursor.execute('SELECT * FROM pubali_replace ORDER BY "S/N" ASC;')
        rows = cursor.fetchall()
        return {"data": rows}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# 2️⃣ Edit Record
# ================================
@router.put("/edit/{sn}", status_code=status.HTTP_200_OK)
def edit_pubali_replace(sn: int, payload: Dict[str, Any] = {}, current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["admin", "superadmin"])
    try:
        set_clause = ", ".join([f'"{k}" = %s' for k in payload.keys()])
        values = tuple(payload.values()) + (sn,)
        cursor.execute(f'UPDATE pubali_replace SET {set_clause} WHERE "S/N" = %s;', values)
        conn.commit()

        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="Edit Pubali Replace",
            target_table="pubali_replace",
            description=f"Edited record S/N={sn}"
        )
        return {"message": f"✅ Record S/N {sn} updated successfully."}
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# 3️⃣ Delete Record
# ================================
@router.delete("/delete/{sn}", status_code=status.HTTP_200_OK)
def delete_pubali_replace(sn: int, current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["admin", "superadmin"])
    try:
        cursor.execute('DELETE FROM pubali_replace WHERE "S/N" = %s;', (sn,))
        conn.commit()

        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="Delete Pubali Replace",
            target_table="pubali_replace",
            description=f"Deleted record S/N={sn}"
        )
        return {"message": f"🗑️ Record S/N {sn} deleted successfully."}
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# 4️⃣ Upload Excel (Bulk Insert/Update)
# ================================
@router.post("/upload", status_code=status.HTTP_201_CREATED)
def upload_pubali_replace_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["admin", "superadmin"])
    try:
        df = pd.read_excel(BytesIO(file.file.read()), sheet_name=2)  # #rd Sheet
        df.columns = [c.strip().upper() for c in df.columns]

        if "POS S/N" not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'POS S/N' column in Excel file.")

        df = df.fillna("").dropna(how="all")

        db_cols = {
            "S/N", "CONFIGURATION DATE", "TID", "MID", "MERCHANT NAME",
            "DBA NAME", "ADDRESS", "BANK CONT. PERSON NAME", "BANK CONT. PERSON NUMBER",
            "DISTRICT", "DIVISION", "ZONE", "POS S/N", "POS BRAND & MODEL",
            "OPERATOR NAME", "SIM NUMBER", "SIM IP", "HOST IP", "HOST PORT",
            "EMI", "VENDOR NAME", "EMI TENURE", "REMARKS", "APPS VERSION",
            "CONFIGURE BY", "INSTALLATION DATE", "INSTALLATION BY",
            "MERCHANT NUMBER", "FIRMWARE VERSION", "QR"
        }

        inserted = updated = skipped = failed = 0

        for _, row in df.iterrows():
            try:
                record = {col: _truncate(row[col]) for col in df.columns if col in db_cols}
                pos_sn = record.get("POS S/N")
                if not pos_sn:
                    skipped += 1
                    continue

                cursor.execute('SELECT "S/N" FROM pubali_replace WHERE "POS S/N" = %s;', (pos_sn,))
                exists = cursor.fetchone()

                if exists:
                    set_clause = ", ".join([f'"{k}" = %s' for k in record])
                    cursor.execute(f'UPDATE pubali_replace SET {set_clause} WHERE "POS S/N" = %s;', (*record.values(), pos_sn))
                    updated += 1
                else:
                    placeholders = ", ".join(["%s"] * len(record))
                    columns = ", ".join([f'"{k}"' for k in record.keys()])
                    cursor.execute(f'INSERT INTO pubali_replace ({columns}) VALUES ({placeholders});', tuple(record.values()))
                    inserted += 1
            except Exception as e:
                traceback.print_exc()
                conn.rollback()
                failed += 1

        conn.commit()

        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="Bulk Upload Pubali Replace",
            target_table="pubali_replace",
            description=f"Inserted={inserted}, Updated={updated}, Skipped={skipped}, Failed={failed}"
        )

        return {"message": "✅ Upload complete", "summary": {"inserted": inserted, "updated": updated, "skipped": skipped, "failed": failed}}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        file.file.close()

# ================================
# 5️⃣ Export Excel
# ================================
@router.get("/download", status_code=status.HTTP_200_OK)
def download_pubali_replace(current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["support", "admin", "superadmin"])
    try:
        columns = [
            "S/N", "CONFIGURATION DATE", "TID", "MID", "MERCHANT NAME",
            "DBA NAME", "ADDRESS", "BANK CONT. PERSON NAME", "BANK CONT. PERSON NUMBER",
            "DISTRICT", "DIVISION", "ZONE", "POS S/N", "POS BRAND & MODEL",
            "OPERATOR NAME", "SIM NUMBER", "SIM IP", "HOST IP", "HOST PORT",
            "EMI", "VENDOR NAME", "EMI TENURE", "REMARKS", "APPS VERSION",
            "CONFIGURE BY", "INSTALLATION DATE", "INSTALLATION BY",
            "MERCHANT NUMBER", "FIRMWARE VERSION", "QR"
        ]
        quoted = [f'"{c}"' for c in columns]
        cursor.execute(f'SELECT {", ".join(quoted)} FROM pubali_replace ORDER BY "S/N";')
        rows = cursor.fetchall()

        df = pd.DataFrame(rows, columns=columns)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet = "Pubali_Replace_Data"
            df.to_excel(writer, index=False, sheet_name=sheet)

            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            worksheet = writer.sheets[sheet]
            for i, col in enumerate(df.columns, 1):
                max_len = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)
                for cell in worksheet[get_column_letter(i)]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        output.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="pubali_replace_data.xlsx"'}
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {e}")

# ================================
# 6️⃣ Download Template
# ================================
@router.get("/template")
def download_pubali_replace_template():
    cols = [
        "S/N", "CONFIGURATION DATE", "TID", "MID", "MERCHANT NAME",
        "DBA NAME", "ADDRESS", "BANK CONT. PERSON NAME", "BANK CONT. PERSON NUMBER",
        "DISTRICT", "DIVISION", "ZONE", "POS S/N", "POS BRAND & MODEL",
        "OPERATOR NAME", "SIM NUMBER", "SIM IP", "HOST IP", "HOST PORT",
        "EMI", "VENDOR NAME", "EMI TENURE", "REMARKS", "APPS VERSION",
        "CONFIGURE BY", "INSTALLATION DATE", "INSTALLATION BY",
        "MERCHANT NUMBER", "FIRMWARE VERSION", "QR"
    ]
    df = pd.DataFrame(columns=cols)
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="pubali_replace_template.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
