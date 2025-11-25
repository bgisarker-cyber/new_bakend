import datetime
from io import BytesIO
from typing import Dict, Any
from fastapi import APIRouter, File, UploadFile, HTTPException, status
from fastapi.responses import StreamingResponse
import pandas as pd
from app.database import conn, cursor
import traceback


router = APIRouter()

# ---------- helpers ----------
#def _truncate(value: Any, max_len: int = 255):
 #   if value is None or str(value).strip() == "":
 #       return None
 #   return str(value).strip()[:max_len]

def _truncate(value: Any, max_len: int = 255):
    if value is None or str(value).strip() == "":
        return r""          # <-- new: store literal N\A instead of NULL
    s = str(value).strip()
    return s[:max_len]

# ---------- 1. get all ----------
@router.get("/islami/all", status_code=status.HTTP_200_OK)
def get_all_islami():
    try:
        cursor.execute("SELECT * FROM Islami ORDER BY id ASC;")
        rows = cursor.fetchall()
        return {"data": rows}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))




# ---------- 4. upload ----------
@router.post("/islami/upload", status_code=status.HTTP_201_CREATED)
def upload_islami_excel(file: UploadFile = File(...)):
    try:
        df = pd.read_excel(BytesIO(file.file.read()))
        df.columns = [c.strip().upper() for c in df.columns]

        if "POS_SERIAL" not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'POS_SERIAL' column in Excel")

        df = df.fillna("").dropna(how="all")

        # normalise column names -> lower case
        db_cols = {
            c.name for c in cursor.description
        } if cursor.description else {
            "configdate", "old_mid", "old_tid", "mid", "tid",
            "merchant_signboard", "address",
            "city", "area", "branch", "vendor", "pos_model", "pos_serial",
            "battery_serial", "old_pos_serial", "sim_operator", "sim_serial_number",
            "ip_address", "port_number", "special_functionality", "tl", "aro",
            "merchant_contact_person", "merchant_contact_number",
            "installation_date", "installation_engineer_name", "handover_to",
            "handover_date", "app_version", "app_release_date", "firmware_version", "remarks"
        }

        inserted = updated = skipped = failed = 0

        for _, row in df.iterrows():
            try:
                record = {}
                for col in df.columns:
                    lc = col.lower()
                    if lc in db_cols:
                        record[lc] = _truncate(row[col])

                pos_serial = record.get("pos_serial")
                if not pos_serial:
                    skipped += 1
                    continue

                cursor.execute("SELECT id FROM Islami WHERE pos_serial = %s;", (pos_serial,))
                exists = cursor.fetchone()

                if exists:
                    set_clause = ", ".join([f"{k} = %s" for k in record])
                    cursor.execute(
                        f"UPDATE Islami SET {set_clause}, update_time = NOW() WHERE pos_serial = %s;",
                        (*record.values(), pos_serial)
                    )
                    updated += 1
                else:
                    placeholders = ", ".join(["%s"] * len(record))
                    cursor.execute(
                        f"INSERT INTO Islami ({', '.join(record.keys())}, create_time, update_time) "
                        f"VALUES ({placeholders}, NOW(), NOW());",
                        tuple(record.values())
                    )
                    inserted += 1
            except Exception as e:
                traceback.print_exc()
                conn.rollback()
                failed += 1

        conn.commit()
        return {
            "message": "✅ Upload completed",
            "summary": {"inserted": inserted, "updated": updated, "failed": failed, "skipped": skipped}
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        file.file.close()

# ---------- 5. template ----------
@router.get("/islami/template")
def download_islami_template():
    cols = [
         "configdate", "old_mid", "old_tid", "mid", "tid",
            "merchant_signboard", "address",
            "city", "area", "branch", "vendor", "pos_model", "pos_serial",
            "battery_serial", "old_pos_serial", "sim_operator", "sim_serial_number",
            "ip_address", "port_number", "special_functionality", "tl", "aro",
            "merchant_contact_person", "merchant_contact_number",
            "installation_date", "installation_engineer_name", "handover_to",
            "handover_date", "app_version", "app_release_date", "firmware_version", "remarks"
    ]
    df = pd.DataFrame(columns=cols)
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="islami_template.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


# Download with Excel
# Download with Excel
@router.get("/islami/download", status_code=status.HTTP_200_OK)
def download_all_islami_records():
    """
    Export all Islami POS records to Excel,
    keeping the same column order as the upload template,
    auto-fitting columns and formatting dates as DD-MM-YYYY.
    """
    try:
        # Define columns in same order as your upload template
        columns = [
            "id", "configdate", "old_mid", "old_tid", "mid", "tid",
            "merchant_signboard", "address",
            "city", "area", "branch", "vendor", "pos_model", "pos_serial",
            "battery_serial", "old_pos_serial", "sim_operator", "sim_serial_number",
            "ip_address", "port_number", "special_functionality", "tl", "aro",
            "merchant_contact_person", "merchant_contact_number",
            "installation_date", "installation_engineer_name", "handover_to",
            "handover_date", "app_version", "app_release_date", "firmware_version",
            "remarks", "create_time"
        ]

        # Fetch all records from DB
        cursor.execute(f"SELECT {', '.join([c.lower() for c in columns])} FROM islami ORDER BY id;")
        rows = cursor.fetchall()

        # Convert to DataFrame
        df = pd.DataFrame(rows, columns=columns)

        # Helper to format dates
        

        # Create Excel in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_name = "Islami_POS_Data"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # Access workbook and sheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Auto-fit column widths and enable text wrapping
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            for i, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                # Limit to avoid extremely wide columns
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width

                # Apply text wrapping to potentially long fields
                if col.lower() in ["address", "remarks", "merchant_signboard"]:
                    for cell in worksheet[get_column_letter(i)]:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        output.seek(0)

        headers = {
            "Content-Disposition": 'attachment; filename="islami_pos_data.xlsx"'
        }

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {e}")

    


