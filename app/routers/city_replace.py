# app/routers/city_replace.py
import datetime
from io import BytesIO
from typing import Dict, Any
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, status
from fastapi.responses import StreamingResponse
import pandas as pd
from app.database import conn, cursor
import traceback
from app.utils.auth import get_current_user
from app.utils.logger import log_user_action
from app.utils.role_check import check_role

router = APIRouter(prefix="/city_replace", tags=["city_replace"])

# ------------------------------------------------------------------
# Helper
# ------------------------------------------------------------------
def _truncate(value: Any, max_len: int = 255):
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip()[:max_len]

# ------------------------------------------------------------------
# 1️⃣ Get All
# ------------------------------------------------------------------
@router.get("/all", status_code=status.HTTP_200_OK)
def get_all_city_replace(current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["support", "admin", "superadmin"])
    try:
        cursor.execute('SELECT * FROM city_replace ORDER BY "SL" ASC;')
        rows = cursor.fetchall()
        return {"data": rows}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------------------------------------------------
# 2️⃣ Edit Row
# ------------------------------------------------------------------
@router.put("/edit/{sl}", status_code=status.HTTP_200_OK)
def edit_city_replace(sl: int, payload: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["admin", "superadmin"])
    try:
        set_clause = ", ".join([f'"{k}" = %s' for k in payload])
        values = tuple(payload.values()) + (sl,)
        cursor.execute(f'UPDATE city_replace SET {set_clause} WHERE "SL" = %s;', values)
        conn.commit()

        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="Edit City Replace",
            target_table="city_replace",
            description=f"Edited SL={sl}"
        )
        return {"message": f"✅ Record SL {sl} updated."}
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------------------------------------------------
# 3️⃣ Delete Row
# ------------------------------------------------------------------
@router.delete("/delete/{sl}", status_code=status.HTTP_200_OK)
def delete_city_replace(sl: int, current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["admin", "superadmin"])
    try:
        cursor.execute('DELETE FROM city_replace WHERE "SL" = %s;', (sl,))
        conn.commit()

        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="Delete City Replace",
            target_table="city_replace",
            description=f"Deleted SL={sl}"
        )
        return {"message": f"🗑️ Record SL {sl} deleted."}
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
# ------------------------------------------------------------------
# 6️⃣ Upload Excel – 6th sheet (index 5), upsert on column F ("POS SERIAL")
# ------------------------------------------------------------------
@router.post("/upload", status_code=status.HTTP_201_CREATED)
def upload_city_replace_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["admin", "superadmin"])
    try:
        df = pd.read_excel(BytesIO(file.file.read()), sheet_name=5)   # 6th sheet
        df.columns = [c.strip() for c in df.columns]

        if "POS SERIAL" not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'POS SERIAL' column.")

        df = df.fillna("").dropna(how="all")

        db_cols = {
            "SL","PURPOSE", "CONFIGDATE", "OLD MID", "OLD TID", "MID", "TID",
            "MERCHANT NAME", "DBA NAME", "ADDRESS", "CITY", "LOCATION",
            "VENDOR", "POS TYPE", "POS MODEL", "POS SERIAL", "OLD POS SERIAL",
            "SIM SERIAL NUMBER", "SIM NUMBER", "SIM CARRIER", "IP ADDRESS",
            "PORT NUMBER", "REASON", "SPECIAL FUNCTIONALITY", "INSTALLATION DATE",
            "MERCHAN CONTACT PERSON", "MERCHAN CONTACT NUMBER", "HANDOVER TO",
            "HANDOVER DATE", "ROLL OUT BY", "ROLL OUT DATE", "APP RELEASE DATE", "NFC"
        }

        inserted = updated = skipped = failed = 0

        for _, row in df.iterrows():
            try:
                record = {}
                for col in df.columns:
                    if col in db_cols:
                        val = row[col]
                        # cast numeric columns
                        if col in {"OLD MID", "OLD TID", "MID", "TID", "PORT NUMBER"}:
                            try:
                                val = int(float(val)) if val != "" else None
                            except Exception:
                                val = None
                        else:
                            val = _truncate(val)
                        record[col] = val

                pos_serial = record.get("POS SERIAL")
                if not pos_serial:
                    skipped += 1
                    continue

                cursor.execute('SELECT "SL" FROM city_replace WHERE "POS SERIAL" = %s;', (pos_serial,))
                exists = cursor.fetchone()

                if exists:
                    set_clause = ", ".join([f'"{k}" = %s' for k in record])
                    cursor.execute(f'UPDATE city_replace SET {set_clause} WHERE "POS SERIAL" = %s;', (*record.values(), pos_serial))
                    updated += 1
                else:
                    placeholders = ", ".join(["%s"] * len(record))
                    columns = ", ".join([f'"{k}"' for k in record])
                    cursor.execute(f'INSERT INTO city_replace ({columns}) VALUES ({placeholders});', tuple(record.values()))
                    inserted += 1
            except Exception as e:
                traceback.print_exc()
                conn.rollback()
                failed += 1

        conn.commit()

        log_user_action(
            user_email=current_user["email"],
            user_role=current_user["role"],
            action="Bulk Upload City Replace",
            target_table="city_replace",
            description=f"Ins={inserted} Upd={updated} Skip={skipped} Fail={failed}"
        )
        return {"message": "✅ Upload complete", "summary": {"inserted": inserted, "updated": updated, "skipped": skipped, "failed": failed}}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        file.file.close()

# ------------------------------------------------------------------
# 5️⃣ Export Excel
# ------------------------------------------------------------------
@router.get("/download")
def download_city_replace(current_user: dict = Depends(get_current_user)):
    check_role(current_user, ["support", "admin", "superadmin"])
    try:
        columns = [
            "SL", "PURPOSE", "CONFIGDATE", "OLD MID", "OLD TID", "MID", "TID",
            "MERCHANT NAME", "DBA NAME", "ADDRESS", "CITY", "LOCATION",
            "VENDOR", "POS TYPE", "POS MODEL", "POS SERIAL", "OLD POS SERIAL",
            "SIM SERIAL NUMBER", "SIM NUMBER", "SIM CARRIER", "IP ADDRESS",
            "PORT NUMBER", "REASON", "SPECIAL FUNCTIONALITY", "INSTALLATION DATE",
            "MERCHAN CONTACT PERSON", "MERCHAN CONTACT NUMBER", "HANDOVER TO",
            "HANDOVER DATE", "ROLL OUT BY", "ROLL OUT DATE", "APP RELEASE DATE", "NFC"
        ]
        quoted = [f'"{c}"' for c in columns]
        cursor.execute(f'SELECT {", ".join(quoted)} FROM city_replace ORDER BY "SL";')
        rows = cursor.fetchall()

        df = pd.DataFrame(rows, columns=columns)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet = "City_Replace_Data"
            df.to_excel(writer, index=False, sheet_name=sheet)

            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            ws = writer.sheets[sheet]
            for i, col in enumerate(df.columns, 1):
                max_len = max(df[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)
                for cell in ws[get_column_letter(i)]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        output.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="city_replace_data.xlsx"'}
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {e}")

# ------------------------------------------------------------------
# 6️⃣ Template Download
# ------------------------------------------------------------------
@router.get("/template")
def download_city_replace_template():
    cols = [
        "SL","PURPOSE", "CONFIGDATE", "OLD MID", "OLD TID", "MID", "TID",
        "MERCHANT NAME", "DBA NAME", "ADDRESS", "CITY", "LOCATION",
        "VENDOR", "POS TYPE", "POS MODEL", "POS SERIAL", "OLD POS SERIAL",
        "SIM SERIAL NUMBER", "SIM NUMBER", "SIM CARRIER", "IP ADDRESS",
        "PORT NUMBER", "REASON", "SPECIAL FUNCTIONALITY", "INSTALLATION DATE",
        "MERCHAN CONTACT PERSON", "MERCHAN CONTACT NUMBER", "HANDOVER TO",
        "HANDOVER DATE", "ROLL OUT BY", "ROLL OUT DATE", "APP RELEASE DATE", "NFC"
    ]
    df = pd.DataFrame(columns=cols)
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="city_replace_template.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)